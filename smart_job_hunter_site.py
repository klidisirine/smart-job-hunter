# smart_job_hunter_site.py
import streamlit as st
import os
import PyPDF2
import json
import time
import io
import openpyxl
import pickle
import re
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import html
from bs4 import BeautifulSoup

# ============================================================
# CONFIGURATION - Sans dotenv
# ============================================================

# Pour Streamlit Cloud : utiliser les secrets
# Pour le test local : utiliser variable d'environnement
try:
    # Essayer d'abord les secrets Streamlit Cloud
    GROQ_API_KEY = st.secrets["GROQ_API_KEY"]
except:
    # En local, utiliser la variable d'environnement
    GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

PROFIL_FILE = "profil_sauvegarde.pkl"

if not GROQ_API_KEY:
    st.error("❌ Clé API GROQ manquante. Configure GROQ_API_KEY dans .env ou les secrets Streamlit")
    st.stop()

# ============================================================
# INITIALISATION SESSION STATE
# ============================================================

def init_session_state():
    """Initialise les variables de session"""
    if 'emails_generes' not in st.session_state:
        st.session_state.emails_generes = {}
    if 'offres_cache' not in st.session_state:
        st.session_state.offres_cache = []
    if 'recherche_effectuee' not in st.session_state:
        st.session_state.recherche_effectuee = False

# ============================================================
# STYLES CSS
# ============================================================

st.markdown("""
<style>
    .main { background-color: #f0f2f6; }
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
    }
    .profile-card {
        background: white;
        border-radius: 12px;
        padding: 1rem;
        margin-bottom: 1rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    .score-badge {
        display: inline-block;
        padding: 4px 12px;
        border-radius: 20px;
        font-weight: bold;
        font-size: 14px;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================
# FONCTIONS
# ============================================================

def clean_html(text):
    """Supprime les balises HTML et nettoie le texte"""
    if not text:
        return ""
    soup = BeautifulSoup(text, 'html.parser')
    text = soup.get_text(separator=' ')
    text = re.sub(r'\s+', ' ', text)
    text = html.unescape(text)
    return text.strip()

def sauvegarder_profil(profil, cv_name):
    with open(PROFIL_FILE, 'wb') as f:
        pickle.dump({"profil": profil, "cv_name": cv_name, "date": datetime.now().isoformat()}, f)

def charger_profil():
    if os.path.exists(PROFIL_FILE):
        with open(PROFIL_FILE, 'rb') as f:
            return pickle.load(f)["profil"]
    return None

def supprimer_profil():
    if os.path.exists(PROFIL_FILE):
        os.remove(PROFIL_FILE)

def extraire_competences_intelligentes(texte_cv):
    """Extraction intelligente des compétences"""
    texte = texte_cv.lower()
    
    competences = {
        "langages": [],
        "frameworks": [],
        "bases_donnees": [],
        "outils": []
    }
    
    langages_list = {
        "python": "Python", "java": "Java", "javascript": "JavaScript",
        "php": "PHP", "c++": "C++", "c#": "C#", "ruby": "Ruby",
        "go": "Go", "rust": "Rust", "typescript": "TypeScript",
        "swift": "Swift", "kotlin": "Kotlin", "html": "HTML",
        "css": "CSS", "c": "C"
    }
    
    frameworks_list = {
        "symfony": "Symfony", "laravel": "Laravel", "django": "Django",
        "flask": "Flask", "react": "React", "angular": "Angular",
        "vue": "Vue.js", "spring": "Spring Boot", "qt": "Qt",
        "javafx": "JavaFX", "flutter": "Flutter", "node": "Node.js"
    }
    
    bdd_list = {
        "mysql": "MySQL", "postgresql": "PostgreSQL", "mongodb": "MongoDB",
        "oracle": "Oracle", "sqlite": "SQLite", "firebase": "Firebase"
    }
    
    outils_list = {
        "docker": "Docker", "git": "Git", "kubernetes": "Kubernetes",
        "linux": "Linux", "vmware": "VMware", "gns3": "GNS3",
        "solidworks": "SolidWorks", "autocad": "AutoCAD", "jenkins": "Jenkins",
        "aws": "AWS", "azure": "Azure", "gcp": "GCP"
    }
    
    for key, value in langages_list.items():
        if re.search(r'\b' + key + r'\b', texte):
            competences["langages"].append(value)
    
    for key, value in frameworks_list.items():
        if re.search(r'\b' + key + r'\b', texte):
            competences["frameworks"].append(value)
    
    for key, value in bdd_list.items():
        if re.search(r'\b' + key + r'\b', texte):
            competences["bases_donnees"].append(value)
    
    for key, value in outils_list.items():
        if re.search(r'\b' + key + r'\b', texte):
            competences["outils"].append(value)
    
    return competences

def extraire_profil_intelligent(texte_cv):
    """Extraction complète du profil"""
    texte = texte_cv.lower()
    
    profil = {
        "nom": "",
        "email": "",
        "ecole": "",
        "niveau": "",
        "domaine": "",
        "langages": [],
        "frameworks": [],
        "bases_donnees": [],
        "outils": [],
        "mots_cles_recherche": []
    }
    
    # Email
    email_match = re.search(r'[\w\.-]+@[\w\.-]+', texte_cv)
    if email_match:
        profil["email"] = email_match.group(0)
    
    # Nom
    for line in texte_cv.split('\n')[:15]:
        line = line.strip()
        if line and 2 <= len(line.split()) <= 4:
            if not any(x in line.lower() for x in ['@', 'tel', '+216', 'linkedin', 'github']):
                if any(w[0].isupper() for w in line.split() if w):
                    profil["nom"] = line
                    break
    
    # École
    ecoles = {
        "esprit": "ESPRIT", "iset": "ISET", "enis": "ENIS", "supcom": "SUP'COM",
        "fst": "FST", "université": "Université", "polytechnique": "Polytechnique"
    }
    for key, value in ecoles.items():
        if key in texte:
            profil["ecole"] = value
            break
    
    # Niveau
    if any(x in texte for x in ["cycle ingenieur", "cycle ingénieur", "5ème", "5eme"]):
        profil["niveau"] = "Cycle Ingénieur"
    elif any(x in texte for x in ["master", "m2"]):
        profil["niveau"] = "Master"
    elif any(x in texte for x in ["licence", "license", "3ème", "3eme"]):
        profil["niveau"] = "Licence"
    elif any(x in texte for x in ["preparatoire", "prépa"]):
        profil["niveau"] = "Cycle Préparatoire"
    
    # Domaine
    if any(x in texte for x in ["mecanique", "mécanique", "cnc", "usinage", "solidworks", "fabrication"]):
        profil["domaine"] = "Génie Mécanique"
        profil["mots_cles_recherche"] = ["stage mecanique", "stage industriel", "genie mecanique"]
    elif any(x in texte for x in ["informatique", "programmation", "developpement", "développement"]):
        profil["domaine"] = "Informatique"
        profil["mots_cles_recherche"] = ["stage informatique", "developpement web", "fullstack"]
    elif any(x in texte for x in ["electrique", "electronique", "circuit"]):
        profil["domaine"] = "Génie Électrique"
        profil["mots_cles_recherche"] = ["stage electrique", "stage electronique"]
    else:
        profil["domaine"] = "Autre"
        profil["mots_cles_recherche"] = ["stage", "internship"]
    
    # Compétences
    competences = extraire_competences_intelligentes(texte_cv)
    profil["langages"] = competences["langages"]
    profil["frameworks"] = competences["frameworks"]
    profil["bases_donnees"] = competences["bases_donnees"]
    profil["outils"] = competences["outils"]
    
    return profil

def chercher_offres_remoteok(mots_cles):
    """Récupère et nettoie les offres RemoteOK"""
    offres = []
    try:
        resp = requests.get("https://remoteok.com/api", headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
        data = resp.json()
        jobs = data[1:] if isinstance(data, list) and len(data) > 1 else []
        
        for job in jobs:
            titre = job.get("position", "") or job.get("title", "")
            description_raw = job.get("description", "")
            description = clean_html(description_raw)
            
            offres.append({
                "titre": titre,
                "entreprise": job.get("company", ""),
                "lieu": job.get("location", "Remote"),
                "description": description[:600],
                "url": job.get("url", job.get("link", "")),
                "date": job.get("date", "Nouveau"),
                "source": "RemoteOK"
            })
        return offres
    except Exception as e:
        return []

def analyser_correspondance(profil, offre):
    """Analyse la correspondance"""
    domaine = profil.get("domaine", "").lower()
    titre = offre.get("titre", "").lower()
    description = offre.get("description", "").lower()
    texte = titre + " " + description
    
    competences_user = profil.get("langages", []) + profil.get("frameworks", [])
    
    points_forts = []
    score = 40
    
    for comp in competences_user[:5]:
        if comp.lower() in texte:
            points_forts.append(comp)
            score += 15
    
    if domaine == "genie mecanique":
        mots_cibles = ["mechanical", "manufacturing", "cnc", "solidworks", "cad", "design", "production", "engineer"]
        for mot in mots_cibles:
            if mot in texte:
                score += 8
                if mot not in points_forts:
                    points_forts.append(mot)
    elif domaine == "informatique":
        mots_cibles = ["software", "developer", "web", "fullstack", "python", "java", "php", "javascript", "engineer"]
        for mot in mots_cibles:
            if mot in texte:
                score += 8
                if mot not in points_forts:
                    points_forts.append(mot)
    
    score = min(score, 95)
    
    if score >= 80:
        conseil = "🎯 Excellente correspondance ! Postulez rapidement."
    elif score >= 60:
        conseil = "👍 Bonne correspondance, personnalisez votre lettre de motivation."
    else:
        conseil = "📚 Correspondance partielle, mettez en avant vos compétences clés."
    
    return {
        "score": score,
        "points_forts": points_forts[:4],
        "conseil": conseil
    }

def generer_email(profil, offre):
    """Génère un email personnalisé selon le profil de l'utilisateur"""
    nom = profil.get("nom", "Candidat")
    email_user = profil.get("email", "")
    titre = offre.get("titre", "le poste")
    entreprise = offre.get("entreprise", "votre entreprise")
    
    # Récupérer les informations du profil
    niveau = profil.get("niveau", "")
    ecole = profil.get("ecole", "")
    domaine = profil.get("domaine", "")
    
    # Récupérer les compétences
    competences = profil.get("langages", []) + profil.get("frameworks", [])
    competences_str = ", ".join(competences[:4]) if competences else "mes compétences techniques"
    
    # Construire la phrase de présentation selon les informations disponibles
    if niveau and ecole:
        presentation = f"Actuellement {niveau} à {ecole}"
    elif ecole:
        presentation = f"Étudiant à {ecole}"
    elif niveau:
        presentation = f"Titulaire d'un {niveau}"
    else:
        presentation = "Je suis étudiant"
    
    # Ajouter le domaine si disponible et pertinent
    if domaine and domaine != "Autre" and domaine != "Non détecté":
        presentation += f" en {domaine}"
    
    # Construction de l'email
    email = f"""Objet : Candidature pour le poste de {titre}

Madame, Monsieur,

{presentation}, je suis vivement intéressé par le poste de {titre} au sein de {entreprise}.

Mon parcours m'a permis d'acquérir des compétences solides en {competences_str}.

Je serais ravi de vous rencontrer pour vous exposer plus en détail ma motivation et mes compétences.

Dans l'attente de votre retour, je vous prie d'agréer, Madame, Monsieur, l'expression de mes salutations distinguées.

{nom}
{email_user if email_user else ''}"""
    
    return email

def exporter_excel(offres_analysees, profil):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offres"
    
    headers = ["Titre", "Entreprise", "Lieu", "Source", "Score", "Points forts", "Conseil", "URL", "Email"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    
    for i, item in enumerate(offres_analysees, start=2):
        offre = item["offre"]
        analyse = item["analyse"]
        
        ws.append([
            offre.get("titre", ""),
            offre.get("entreprise", ""),
            offre.get("lieu", ""),
            offre.get("source", ""),
            f"{analyse.get('score', 0)}%",
            ", ".join(analyse.get("points_forts", [])),
            analyse.get("conseil", ""),
            offre.get("url", ""),
            item.get("email", "")
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# INTERFACE STREAMLIT
# ============================================================

def main():
    st.set_page_config(page_title="Smart Job Hunter Pro", page_icon="🎯", layout="wide")
    
    # Initialiser session state
    init_session_state()
    
    st.markdown("""
    <div class="main-header">
        <h1>🎯 Smart Job Hunter Pro</h1>
        <p>Trouve le stage qui te correspond vraiment</p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")
        st.info("📡 Source: RemoteOK (offres internationales)")
        st.success("🎯 100% Gratuit")
        st.markdown("---")
        if st.button("🗑️ Nouveau CV", use_container_width=True):
            supprimer_profil()
            st.session_state.emails_generes = {}
            st.session_state.offres_cache = []
            st.session_state.recherche_effectuee = False
            st.rerun()
    
    profil = charger_profil()
    
    if not profil:
        st.markdown("### 📄 Étape 1 : Dépose ton CV")
        cv_file = st.file_uploader("Format PDF uniquement", type=["pdf"])
        
        if cv_file:
            with st.spinner("📖 Analyse intelligente de ton CV..."):
                reader = PyPDF2.PdfReader(cv_file)
                texte = ""
                for page in reader.pages:
                    texte += page.extract_text()
                
                profil = extraire_profil_intelligent(texte)
                if profil:
                    sauvegarder_profil(profil, cv_file.name)
                    st.success("✅ CV analysé avec succès !")
                    st.rerun()
                else:
                    st.error("❌ Erreur lors de l'analyse")
        return
    
    # Affichage profil
    st.markdown("### 👤 Ton profil")
    
    col1, col2 = st.columns([2, 1])
    with col1:
        st.markdown(f"""
        <div class="profile-card">
            <strong>📛 Nom :</strong> {profil.get('nom', 'Non détecté')}<br>
            <strong>🎓 École :</strong> {profil.get('ecole', 'Non détectée')}<br>
            <strong>📚 Niveau :</strong> {profil.get('niveau', 'Non détecté')}<br>
            <strong>🔧 Domaine :</strong> {profil.get('domaine', 'Non détecté')}
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        mots_suggestions = profil.get("mots_cles_recherche", ["stage"])
        st.markdown(f"""
        <div class="profile-card">
            <strong>💡 Mots-clés suggérés :</strong><br>
            <code>{' | '.join(mots_suggestions[:4])}</code>
        </div>
        """, unsafe_allow_html=True)
    
    with st.expander("📋 Compétences détectées", expanded=False):
        col_a, col_b = st.columns(2)
        with col_a:
            if profil.get("langages"):
                st.write("**💻 Langages :**", ", ".join(profil["langages"]))
            if profil.get("frameworks"):
                st.write("**⚙️ Frameworks :**", ", ".join(profil["frameworks"]))
        with col_b:
            if profil.get("bases_donnees"):
                st.write("**🗄️ Bases de données :**", ", ".join(profil["bases_donnees"]))
            if profil.get("outils"):
                st.write("**🛠️ Outils :**", ", ".join(profil["outils"]))
    
    st.divider()
    
    # Recherche
    st.markdown("### 🔍 Étape 2 : Recherche d'offres")
    
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        mots_defaut = " ".join(profil.get("mots_cles_recherche", ["stage"])[:3])
        mots_cles = st.text_input("Mots-clés de recherche", value=mots_defaut)
    with col2:
        score_min = st.slider("Score minimum", 0, 80, 40)
    with col3:
        st.markdown(" ")
        st.markdown(" ")
        rechercher = st.button("🚀 Rechercher", type="primary", use_container_width=True)
    
    if rechercher:
        with st.spinner("🔍 Scraping des offres en cours..."):
            offres = chercher_offres_remoteok(mots_cles)
        
        if not offres:
            st.error("❌ Aucune offre trouvée. Essaie d'autres mots-clés.")
            return
        
        st.success(f"✅ {len(offres)} offres trouvées")
        
        offres_analysees = []
        progress = st.progress(0)
        
        for i, offre in enumerate(offres[:40]):
            progress.progress((i + 1) / min(len(offres), 40))
            analyse = analyser_correspondance(profil, offre)
            offres_analysees.append({"offre": offre, "analyse": analyse, "email": ""})
        
        progress.empty()
        
        offres_filtrees = [o for o in offres_analysees if o["analyse"].get("score", 0) >= score_min]
        offres_filtrees.sort(key=lambda x: x["analyse"].get("score", 0), reverse=True)
        
        # Sauvegarder dans session state
        st.session_state.offres_cache = offres_filtrees
        st.session_state.recherche_effectuee = True
        st.session_state.emails_generes = {}  # Réinitialiser les emails
        
        col1, col2, col3 = st.columns(3)
        col1.metric("📊 Total analysé", len(offres[:40]))
        col2.metric("✅ Offres pertinentes", len(offres_filtrees))
        if offres_filtrees:
            col3.metric("🏆 Meilleur score", f"{offres_filtrees[0]['analyse'].get('score', 0)}%")
    
    # Afficher les offres depuis le cache
    if st.session_state.recherche_effectuee and st.session_state.offres_cache:
        for idx, item in enumerate(st.session_state.offres_cache):
            offre = item["offre"]
            analyse = item["analyse"]
            score = analyse.get("score", 0)
            
            if score >= 70:
                badge = "🟢 Excellent"
            elif score >= 50:
                badge = "🟡 Bon"
            else:
                badge = "🟠 À améliorer"
            
            # Créer une clé unique pour cette offre
            offre_key = f"{offre['titre']}_{offre['entreprise']}_{idx}"
            
            with st.expander(f"**{offre['titre']}** — {offre['entreprise']}"):
                col1, col2 = st.columns([3, 1])
                with col1:
                    st.write(f"📍 **Lieu :** {offre.get('lieu', 'Non spécifié')}")
                    st.write(f"📝 **Description :** {offre.get('description', '')[:400]}...")
                    st.write(f"✅ **Points forts :** {', '.join(analyse.get('points_forts', []))}")
                    st.write(f"💡 **Conseil :** {analyse.get('conseil', '')}")
                with col2:
                    st.markdown(f"<span class='score-badge' style='background-color:#28a745;color:white'>{badge} {score}%</span>", unsafe_allow_html=True)
                
                if offre.get("url"):
                    st.link_button("🔗 Voir l'offre originale", offre["url"])
                
                # Bouton générer email avec session state
                if st.button(f"✉️ Générer email", key=f"btn_{idx}_{offre_key}"):
                    with st.spinner("✍️ Génération..."):
                        email = generer_email(profil, offre)
                        st.session_state.emails_generes[offre_key] = email
                        st.rerun()
                
                # Afficher l'email s'il a été généré
                if offre_key in st.session_state.emails_generes:
                    st.text_area("📧 Email prêt à copier :", 
                                st.session_state.emails_generes[offre_key], 
                                height=200, 
                                key=f"email_{idx}_{offre_key}")
        
        # Export Excel
        if st.session_state.offres_cache:
            excel_data = exporter_excel(st.session_state.offres_cache, profil)
            st.download_button(
                label="📥 Télécharger le rapport Excel",
                data=excel_data,
                file_name=f"offres_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()